import os
import fitz  # PyMuPDF
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

output_file = "output.xlsx"

def append_to_csv(pdfname, level, description, count, filename):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['C'].width = 200

        ws.append(['PDF Name', 'Level', 'Description', 'Count'])  # 標題列
        wb.save(filename)

    ws.append([pdfname, level, description, count])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # 加一點空間

    wb.save(filename)

def print_pages_starting_with_summary(pdf_path):
    blHasProblem = False

    try:
        doc = fitz.open(pdf_path)
        for page_num in range(len(doc)):
            text = doc[page_num].get_text()
            lines = text.splitlines()
            blFindOutline = True
            blEnterProblemList = False
            lineIdx = 0
            problemText = ""

            while lineIdx < len(lines):
                line = lines[lineIdx]
                lineIdx += 1

                if not line.strip():  # 空白
                    continue
                if line == "目錄":
                    continue
                if blFindOutline is True:
                    if line != "摘要":
                        break
                    blFindOutline = False

                if line[:1] == "重" or line[:1] == "高" or line[:1] == "中" or line[:1] == "低" or line[:1] == "參":
                    level = line[:1]
                    blEnterProblemList = True
                    blHasProblem = True
                    problemText = line[1:]
                else:
                    if blEnterProblemList is True:
                        break
                    continue

                while 1:
                    if lineIdx >= len(lines):
                        raise Exception("Error: 已到達最後一行，沒有下一行可讀")

                    line = lines[lineIdx].strip()
                    lineIdx += 1

                    if line.isdigit():
                        count = line.strip()
                        break
                    else:
                        problemText += line

                print(f"{level}: {problemText}, count: {count}")
                append_to_csv(os.path.basename(pdf_path), level, problemText, count, output_file)

        if blHasProblem is False:
            append_to_csv(os.path.basename(pdf_path), "-", "無", "0", output_file)


    except Exception as e:
        print(f"❌ 無法處理檔案 {pdf_path}: {e}")

def main(folder_path="."):
    if os.path.exists(output_file):
        os.remove(output_file)

    for file in os.listdir(folder_path):
        if file.lower().endswith(".pdf"):
            full_path = os.path.join(folder_path, file)
            print_pages_starting_with_summary(full_path)

if __name__ == "__main__":
    main("./pdfs")
    #input("\n✅ 執行完畢，按 Enter 鍵關閉視窗...")
